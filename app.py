import streamlit as st
import cv2
import numpy as np
import tempfile
import time
from PIL import Image
from ultralytics import YOLO
from paddleocr import PaddleOCR, draw_ocr
import pandas as pd
import plotly.express as px
import openpyxl
from datetime import datetime
import json
import requests
from streamlit_lottie import st_lottie


# Create and Load data

parking_data = 'parking_data.xlsx'
current_parking_status = 'current_parking_status'
history = 'history'

df1 = pd.DataFrame(columns=['Id', 'Current Plate number'])
df2 = pd.DataFrame(columns=['Id', 'Plate Number', 'Time get in', 'Time get out', 'Parking duration'])

with pd.ExcelWriter(parking_data) as writer:
    df1.to_excel(writer, sheet_name=current_parking_status, index=False)
    df2.to_excel(writer, sheet_name=history, index=False)

workbook = openpyxl.load_workbook(parking_data)



# Load trained model
DEMO_IMAGE = 'demo.jpg'
DEMO_VIDEO = 'demo.mp4'
model_path = 'best.pt'
class_name_dict = {0: 'motorbike plate'}

# Plate Detection
model = YOLO(model_path)

# Plate Recognition
ocr = PaddleOCR(use_angle_cls=True, lang='en')

st.title('Motorbike Licence Plate Detector')

st.markdown(
    """
    <style>
    [data-testid="stSidebar"][aria-expanded="true"] > div:first-child{
        width: 350px
    }
    
    [data-testid="stSidebar"][aria-expanded="false"] > div:first-child{
        width: 350px
        margin-left: -350px
    }
    <style>
    
    """,
    unsafe_allow_html=True,
)

def load_lottiefile(filepath: str):
    with open(filepath, "r") as f:
        return json.load(f)
@st.cache_data()
def image_resize(image, width=None, height=None, inter=cv2.INTER_AREA):
    dim = None
    (h, w) = image.shape[:2]

    if width is None and height is None:
        return image

    if width is None:
        r = width/float(w)
        dim = (int(w*r), height)

    else:
        r = width/float(w)
        dim = (width, int(h*r))

    # resize the image
    resized = cv2.resize(image, dim, interpolation=inter)

    return resized

app_mode = st.sidebar.selectbox('Choose the App mode',
                                ['About App', 'Run on Image', 'Run on Video']
                                )
if app_mode == 'About App':
    st.markdown('In this Application we are using **Trained model** for for detecting motorbike plate')

    st.markdown(
        """
        <style>
        [data-testid="stSidebar"][aria-expanded="true"] > div:first-child{
            width: 350px
        }

        [data-testid="stSidebar"][aria-expanded="false"] > div:first-child{
            width: 350px
            margin-left: -350px
        }
        <style>

        """,
        unsafe_allow_html=True,
    )

    lottie_car = load_lottiefile("car.json")  # replace link to local lottie file
    st_lottie(
        lottie_car,
        speed=1,
        reverse=False,
        loop=True,
        quality="high",  # medium ; high  # canvas
        height=500,
        width=None,
        key=None,
    )

    #st.video()
    st.markdown('''
    # About Us \n
      We are **Le Viet Anh Huy**, **Hoang Tuan Anh**, and **Vo Nguyen Trung Quan** from **DUT**. \n
      
      If you you want to buy our website, do not hesitate to contact us at \n
        Phone Number: 0836013299
        Email: pbl3quamon@gmail.com
    ''')


elif app_mode == 'Run on Image':
    st.sidebar.markdown('---')
    st.markdown(
        """
        <style>
        [data-testid="stSidebar"][aria-expanded="true"] > div:first-child{
            width: 350px
        }

        [data-testid="stSidebar"][aria-expanded="false"] > div:first-child{
            width: 350px
            margin-left: -350px
        }
        <style>

        """,
        unsafe_allow_html=True,
    )

    kpi1, kpi2 = st.columns(2)
    # f"<h1 style='text-align: center; color: red;'>{int(fps)}</h1>",
    #                                     unsafe_allow_html=True
    with kpi1:
        st.markdown(f"<h3 style='text-align: center;'>{'Detected Motorbike Plates'}</h3>", unsafe_allow_html=True)
        kpi1_text = st.markdown("0")

    with kpi2:
        st.markdown(f"<h3 style='text-align: center;'>{'Motorbike Plate Number'}</h3>", unsafe_allow_html=True)
        kpi2_text = st.markdown("0")


    max_plates = st.sidebar.number_input('Maximum Number of Motorbike Plate', value=2, min_value=1)
    st.sidebar.markdown('---')
    detection_confidence = st.sidebar.slider('min Detection Confidence', min_value=0.0, max_value=1.0, value=0.5)
    st.sidebar.markdown('---')

    img_file_buffer = st.sidebar.file_uploader("upload an Image", type=['jpg', 'png', 'jpeg'])
    if img_file_buffer is not None:
        image = np.array(Image.open(img_file_buffer))

    else:
        demo_image = DEMO_IMAGE
        image = np.array(Image.open(demo_image))

    st.sidebar.text('Original Image')
    st.sidebar.image(image)
    plate_count = 0

    ## Dashboard
    threshold = detection_confidence
    results = model(image)[0]
    print(results)
    for result in results.boxes.data.tolist():
        x1, y1, x2, y2, score, class_id = result

        if score > threshold:
            cv2.rectangle(image, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 1)
            cv2.putText(image, str(round(score, 2)), (int(x1), int(y1 - 10)),
                        cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 255, 0), 1, cv2.LINE_AA)
            plate_count += 1
            ocr_img = image[int(y1):int(y2), int(x1):int(x2), :].copy()
            gray_img = cv2.cvtColor(ocr_img, cv2.COLOR_BGR2GRAY)
            bin_img = cv2.adaptiveThreshold(gray_img, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 11, 2)
            bin_img = gray_img
            #_, bin_img = cv2.threshold(gray_img, 128, 255, cv2.THRESH_BINARY)
            st.image(bin_img, use_column_width=True)
            result = ocr.ocr(bin_img, cls=True)
            print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
            print(len(result))
            print(result)
            if result[0] == []:
                plate_number = "Error"
            else:
                plate_number = result[0][0][1][0] + ' ' + result[0][1][1][0]
            kpi1_text.write(f"<h2 style='text-align: center; color: red;'>{plate_count}</h2>", unsafe_allow_html=True)
            kpi2_text.write(f"<h2 style='text-align: center; color: red;'>{plate_number}</h2>", unsafe_allow_html=True)

        current_parking_sheet = workbook[current_parking_status]
        # current_parking_sheet.cell(row=3, column=2, value=plate_number)
        #
        current_parking_sheet['B' + str(plate_count + 1)] = plate_number
        workbook.save(parking_data)
    out_pd = pd.read_excel(parking_data,
                           sheet_name=current_parking_status,
                           usecols='B')

    st.subheader('Output Image')
    st.image(image, use_column_width=True)


    plate_index = 2

    st.dataframe(out_pd, use_container_width=True)
    # LOAD DATAFRAME


elif app_mode == 'Run on Video':

    st.set_option('deprecation.showfileUploaderEncoding', False)

    use_webcam = st.sidebar.button('Use Webcam')
    record = st.sidebar.checkbox('Record Video')

    if record:
        st.checkbox("Recording", value=True)

    st.sidebar.markdown('---')
    st.markdown(
        """
        <style>
        [data-testid="stSidebar"][aria-expanded="true"] > div:first-child{
            width: 350px
        }

        [data-testid="stSidebar"][aria-expanded="false"] > div:first-child{
            width: 350px
            margin-left: -350px
        }
        <style>

        """,
        unsafe_allow_html=True,
    )

    st.subheader('Output Video')
    #kpi1_text = st.markdown("0")

    max_plates = st.sidebar.number_input('Maximum Number of Motorbike Plate', value=2, min_value=1)
    st.sidebar.markdown('---')
    detection_confidence = st.sidebar.slider('min Detection Confidence', min_value=0.0, max_value=1.0, value=0.5)
    tracking_confidence = st.sidebar.slider('min Tracking Confidence', min_value=0.0, max_value=1.0, value=0.5)
    st.sidebar.markdown('---')

    stframe = st.empty()
    stframe1 = st.empty()
    video_file_buffer = st.sidebar.file_uploader("Upload a Video", type=['mp4', 'mov', 'avi', 'm4v', 'asf'])
    tffile = tempfile.NamedTemporaryFile(delete=False)

    ## Input Video
    if not video_file_buffer:
        if use_webcam:
            vid = cv2.VideoCapture(0)
            vid1 = cv2.VideoCapture(1)
        else:
            vid = cv2.VideoCapture(DEMO_VIDEO)
            vid1 = cv2.VideoCapture(DEMO_VIDEO)
            tffile.name = DEMO_VIDEO
    else:
        tffile.write(video_file_buffer.read())
        vid = cv2.VideoCapture(tffile.name)
        vid1 = cv2.VideoCapture(tffile.name)

    width = int(vid.get(cv2.CAP_PROP_FRAME_WIDTH))
    height = int(vid.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps_input = int(vid.get(cv2.CAP_PROP_FPS))

    width1 = int(vid1.get(cv2.CAP_PROP_FRAME_WIDTH))
    height1 = int(vid1.get(cv2.CAP_PROP_FRAME_HEIGHT))
    fps_input1 = int(vid1.get(cv2.CAP_PROP_FPS))

    # Recording Part
    codec = cv2.VideoWriter_fourcc('M', 'J', 'P', 'G')
    out = cv2.VideoWriter('output1.mp4', codec, fps_input, (width, height))

    st.sidebar.text('Input Video Get in')
    st.sidebar.video(tffile.name)

    st.sidebar.text('Input Video Get out')
    st.sidebar.video(tffile.name)

    fps = 0
    i = 0

    fps1 = 0
    i1 = 0

    kpi1, kpi2, kpi3 = st.columns(3)

    with kpi1:
        st.markdown(f"<h5 style='text-align: center;'>{'Frame Rate'}</h5>", unsafe_allow_html=True)
        kpi1_text = st.markdown("0")

    with kpi2:
        st.markdown(f"<h5 style='text-align: center;'>{'Detected Motorbike Plates'}</h5>", unsafe_allow_html=True)
        kpi2_text = st.markdown("0")

    with kpi3:
        st.markdown(f"<h5 style='text-align: center;'>{'Image Width'}</h5>", unsafe_allow_html=True)
        kpi3_text = st.markdown("0")

    st.markdown("<hr/>", unsafe_allow_html=True)

    # Motorbike Plate Predictor
    hide_dataframe_row_index = """
                        <style>
                        .row_heading.level0 {display:none}
                        .blank {display:none}
                        </style>
                        """

    plate_count = 0
    prevTime = 0
    threshold = detection_confidence
    placeholder = st.empty()
    container = st.container()
    plate_count = 0
    plates = []
    plates1 = []
    while vid.isOpened() or vid1.isOpened():
        i += 1
        i1 += 1

        ret, frame = vid.read()
        ret1, frame1 = vid1.read()
        if not ret:
            continue
        if not ret1:
            continue
        #frame = cv2.cvtColor(frame, cv2.COLOR_BGR2RGB)
        frame.flags.writeable = True
        frame1.flags.writeable = True
        #frame = cv2.cvtColor(frame, cv2.COLOR_RGB2BGR)
        results = model(frame)[0]
        results1 = model(frame1)[0]
        for result in results.boxes.data.tolist():
            x1, y1, x2, y2, score, class_id = result

            if score > threshold:
                cv2.rectangle(frame, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 1)
                cv2.putText(frame, str(round(score, 2)), (int(x1), int(y1 - 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 255, 0), 1, cv2.LINE_AA)
                ocr_img = frame[int(y1):int(y2), int(x1):int(x2), :].copy()
                gray_img = cv2.cvtColor(ocr_img, cv2.COLOR_BGR2GRAY)
                bin_img = cv2.adaptiveThreshold(gray_img, 255, cv2.ADAPTIVE_THRESH_MEAN_C, cv2.THRESH_BINARY, 11, 2)
                # _, bin_img = cv2.threshold(gray_img, 128, 255, cv2.THRESH_BINARY)
                bin_img = gray_img
                result = ocr.ocr(bin_img, cls=True)
                plate_count += 1
                print("!!!!!!!!!!!!!!!!!!!!!!!!!!!!!")
                print(result)
                plate_number = ''
                if result[0] == []:
                    plate_number = "Error"
                else:
                    for i in range(len(result[0])):
                     plate_number = plate_number + result[0][i][1][0] + ' '
                current_parking_sheet = workbook[current_parking_status]
                history_sheet = workbook[history]
                #current_parking_sheet.cell(row=3, column=2, value=plate_number)
                if plate_number in plates:
                    print('Plate Number Exist')
                    plate_count -= 1
                else:
                    plates.append(plate_number)
                    currentDateAndTime = datetime.now()
                    currentTime = currentDateAndTime.strftime("%H:%M:%S")
                    history_sheet['A' + str(plate_count + 1)] = str(plate_count + 1)
                    history_sheet['B' + str(plate_count + 1)] = plate_number
                    history_sheet['C' + str(plate_count + 1)] = currentTime
                    current_parking_sheet['B' + str(plate_count + 1)] = plate_number
                    current_parking_sheet['A' + str(plate_count + 1)] = str(plate_count + 1)
                #current_parking_sheet.drop_duplicates(subset=["current_parking_status"], keep='first', inplace=True)
                workbook.save(parking_data)

        #time.sleep(1)
        for result1 in results1.boxes.data.tolist():
            x1, y1, x2, y2, score, class_id = result1

            if score > threshold:
                cv2.rectangle(frame1, (int(x1), int(y1)), (int(x2), int(y2)), (0, 255, 0), 1)
                cv2.putText(frame1, str(round(score, 2)), (int(x1), int(y1 - 10)),
                            cv2.FONT_HERSHEY_SIMPLEX, 1.3, (0, 255, 0), 1, cv2.LINE_AA)

                history_sheet = workbook[history]
                if plate_number in plates1:
                    print('Plate Number Exist')
                    plate_count -= 1
                else:
                    plates.append(plate_number)
                    currentDateAndTime1 = datetime.now()
                    currentTime1 = currentDateAndTime.strftime("%H:%M:%S")
                    history_sheet['D' + str(plate_count + 1)] = currentTime1
                    duration = currentDateAndTime1 - currentDateAndTime
                    duration = duration.seconds
                    #duration1 = duration.strftime("%H:%M:%S")
                    history_sheet['E' + str(plate_count + 1)] = str(duration)

                #current_parking_sheet.drop_duplicates(subset=["current_parking_status"], keep='first', inplace=True)
                workbook.save(parking_data)
        # FPS Counter Logic
        currTime = time.time()
        fps = 1 / (currTime - prevTime)
        prevTime = currTime

        if record:
            out.write(frame)

        # Dashboard
        kpi1_text.write(f"<h2 style='text-align: center; color: red;'>{int(fps)}</h2>",
                                    unsafe_allow_html=True)
        kpi2_text.write(f"<h2 style='text-align: center; color: red;'>{plate_count}</h2>",
                                unsafe_allow_html=True)
        kpi3_text.write(f"<h2 style='text-align: center; color: red;'>{width}</h2>",
                                unsafe_allow_html=True)
        frame = cv2.resize(frame, (0,0), fx=0.8, fy=0.8)
        frame = image_resize(image=frame, width=640)
        stframe.image(frame, channels='BGR', use_column_width=True)
        out_pd = pd.read_excel(parking_data,
                               sheet_name=current_parking_status,
                               usecols='A:B')
        #placeholder = st.empty()
        st.markdown(hide_dataframe_row_index, unsafe_allow_html=True)
        with placeholder.container():
            st.dataframe(out_pd, use_container_width=True)
        # st.stop()
        # st.rerun()
        stframe1.image(frame1, channels='BGR', use_column_width=True)
        #plate_count = 0

        # show_history = st.sidebar.button('Watch History', key=1)
        #
        #
        # if show_history:
        #     st.checkbox("Watch History", value=True)
        # if show_history:
        #     st.markdown("History")
        #     out_pd = pd.read_excel(parking_data,
        #                            sheet_name=history,
        #                            usecols='A:E')



st.sidebar.title('Licence Plate Sidebar')